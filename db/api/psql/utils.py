from openpyxl import Workbook
from psycopg2.extensions import connection

from . import db


def insert_blank_data(patient_id: int, conn: connection):
    cursor = conn.cursor()

    insert_blanks_query = """ INSERT INTO answer (question_id, station_id, answers, patient_id)
                                  SELECT question_id, station_id, '', %s
                                  FROM question
                                  WHERE station_id !=
                                    (SELECT station_id
                                    FROM station
                                    WHERE station_name = 'Registration');  """

    patient_id_to_insert = (patient_id,)
    cursor.execute(insert_blanks_query, patient_id_to_insert)
    conn.commit()

    print("Successful addition of blank answers")
    return "Data successfully inserted"


def insert_blank_registration_data(patient_id: int, conn: connection):
    cursor = conn.cursor()

    insert_blanks_query = """ INSERT INTO answer (question_id, station_id, answers, patient_id)
                                  SELECT question_id, station_id, '', %s
                                  FROM question
                                  WHERE (question_id NOT IN
                                  (SELECT question_id
                                  FROM answer
                                  WHERE (patient_id = %s)))
                                  AND station_id =
                                    (SELECT station_id
                                    FROM station
                                    WHERE station_name = 'Registration')  """

    cursor.execute(
        insert_blanks_query,
        (
            patient_id,
            patient_id,
        ),
    )
    conn.commit()

    print("Successful addition of blank answers")
    return "Data successfully inserted"


def create_station(conn: connection):
    cursor = conn.cursor()

    create_stations_table_query = """CREATE TABLE IF NOT EXISTS station
                  (station_id SERIAL PRIMARY KEY,
                  station_name TEXT UNIQUE,
                  availability BOOLEAN); """
    cursor.execute(create_stations_table_query)
    conn.commit()
    print("Table station created successfully in PostgreSQL ")


def create_patient(conn: connection):
    cursor = conn.cursor()

    create_patients_table_query = """CREATE TABLE IF NOT EXISTS patient
                  (patient_id SERIAL PRIMARY KEY,
                  status BOOLEAN,
                  in_queue_station INTEGER,
                  completed_station INTEGER[]); """
    cursor.execute(create_patients_table_query)
    conn.commit()
    print("Table patient created successfully in PostgreSQL ")


def create_question(conn: connection):
    cursor = conn.cursor()

    create_questions_table_query = """CREATE TABLE IF NOT EXISTS question
                  (question_id SERIAL PRIMARY KEY,
                  question_num INTEGER,
                  question TEXT,
                  required BOOLEAN,
                  options varchar[],
                  station_id INTEGER,
                  type_id INTEGER); """
    cursor.execute(create_questions_table_query)
    conn.commit()
    print("Table question created successfully in PostgreSQL ")


def create_answer(conn: connection):
    cursor = conn.cursor()

    create_answers_table_query = """CREATE TABLE IF NOT EXISTS answer
                  (answer_id SERIAL PRIMARY KEY,
                  patient_id INTEGER,
                  answers TEXT,
                  question_id INTEGER,
                  station_id INTEGER); """
    cursor.execute(create_answers_table_query)
    conn.commit()
    print("Table answer created successfully in PostgreSQL ")


def create_type(conn: connection):
    cursor = conn.cursor()

    create_type_table_query = """CREATE TABLE IF NOT EXISTS type
                  (type_id SERIAL PRIMARY KEY,
                  type_info TEXT); """
    cursor.execute(create_type_table_query)
    conn.commit()
    print("Table type created successfully in PostgreSQL ")


def db_setup():
    with db.getconn() as conn:
        create_station(conn)
        create_patient(conn)
        create_question(conn)
        create_answer(conn)
        create_type(conn)


def insert_patient(status, completed_station, conn: connection):
    cursor = conn.cursor()

    postgres_insert_query = """ INSERT INTO patient (patient_id, status, in_queue_station, completed_station)
                                    VALUES (DEFAULT, %s, -1, %s)"""
    record_to_insert = (
        status,
        completed_station,
    )
    cursor.execute(postgres_insert_query, record_to_insert)
    conn.commit()
    count = cursor.rowcount
    print(count, "Record inserted successfully into patient table")


def insert_answer(patient_id, answer, question_id, station_id, conn: connection):
    cursor = conn.cursor()
    postgres_insert_query = """ INSERT INTO answer (answer_id, patient_id, answers, question_id, station_id) VALUES (DEFAULT, %s, %s, %s, %s)"""
    record_to_insert = (
        patient_id,
        answer,
        question_id,
        station_id,
    )
    cursor.execute(postgres_insert_query, record_to_insert)
    conn.commit()
    count = cursor.rowcount
    print(count, "Record inserted successfully into answer table")


def copy_query_to_file(query: str, file, conn: connection) -> None:
    """Copies results from a query into a file-like object, in csv format."""
    q = f"COPY {query} TO STDOUT WITH CSV HEADER"
    cursor = conn.cursor()
    cursor.copy_expert(q, file)


def to_xlsx(conn: connection):
    """to_xlsx returns a openpyxl Workbook object.

    Each station is its own sheet. The first column of each
    sheet contains the questions at that station. Each subsequent
    column represents a single patient's responses to the
    questions. The patient's column is given by patient_id+1, and
    this is shared across worksheets.
    """
    wb = Workbook()
    cursor = conn.cursor()

    # take all station names first to create the worksheets
    cursor.execute("select distinct station_id, station_name from station;")
    res = cursor.fetchall()
    station_ids = [x[0] for x in res]
    station_names = [x[1] for x in res]
    for sname in station_names:
        wb.create_sheet(sname)

    # now grab the remaining data for each station
    for (sid, sname) in zip(station_ids, station_names):
        cursor.execute(
            f"""
        select q.question_id, q.question, a.answer, a.patient_id
        from question q inner join answer a
        on q.question_id=a.question_id
        where q.station_id={sid}
        order by q.question_id;
        """
        )
        res = cursor.fetchall()
        ws = wb.get_sheet_by_name(sname)

        # create a map from qid to the row num, so that rows
        # are flushed to the top
        qmap = {}
        for (i, qid) in enumerate({x[0] for x in res}):
            qmap[qid] = i + 1

        for (qid, q) in {(x[0], x[1]) for x in res}:
            ws.cell(row=qmap[qid], column=1, value=q)

        for x in res:
            qid, ans, pid = x[0], x[2], x[3]
            ws.cell(row=qmap[qid], column=pid + 1, value=ans)

    return wb
